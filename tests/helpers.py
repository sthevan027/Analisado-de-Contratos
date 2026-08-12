from pathlib import Path

from openpyxl import Workbook


def criar_planilha(path: Path, abas: list[str]) -> Path:
    workbook = Workbook()
    workbook.active.title = abas[0]
    for aba in abas[1:]:
        workbook.create_sheet(aba)
    workbook.save(path)
    return path


def criar_qqp_minima(path: Path) -> Path:
    workbook = Workbook()
    aba = workbook.active
    aba.title = "QQP"
    aba.append(["ITEM", "DESCRIÇÃO", "UNID.", "QUANT.", "PREÇOS (R$)"])
    aba.append([None, None, None, None, "UNITÁRIO", "TOTAL"])
    aba.append(["1", "Obras civis", None, None, None, None])
    aba.append(["1.1", "Fundações", None, None, None, None])
    aba.append(["1.1.1", "Escavação", "m3", 5, 5.00, 25.00])
    aba.append(["1.1.2", "Concreto", "m3", 2, 5.00, 10.00])
    workbook.save(path)
    return path


def criar_dados_minimos(path: Path) -> Path:
    workbook = Workbook()
    aba = workbook.active
    aba.title = "Dados"
    aba.append(["ITEM", "DESCRIÇÃO", "UNIDADE", "QUANTIDADE", "VALOR UNITÁRIO", "VALOR TOTAL", "SEÇÃO"])
    aba.append(["2.1", "Serviço legado", "un", 5, 250.00, 1250.00, "Elétrica"])
    workbook.save(path)
    return path


def criar_qqp_com_cpu(path: Path, total_geral: float = 100) -> Path:
    workbook = Workbook()
    qqp = workbook.active
    qqp.title = "QQP"
    qqp.append(["ITEM", "DESCRIÇÃO", "UNID.", "QUANT.", "PREÇOS (R$)"])
    qqp.append([None, None, None, None, "UNITÁRIO", "TOTAL"])
    qqp.append(["1", "Obras civis", None, None, None, None])
    qqp.append(["1.1", "Fundações", None, None, None, None])
    qqp.append(["1.1.1", "Serviço com CPU", "vb", 1, 100.00, 100.00])
    qqp.append(["1.1.2", "Serviço sem CPU", "vb", 1, 50.00, 50.00])

    cpu = workbook.create_sheet("CPU")
    cpu.append(["ITEM:", "1.1.1", "UNID.:", "vb", 100])
    cpu.append(["ITEM", "DESCRIÇÃO", "UNID", "QUANTIDADE", "VALOR UNIT.", "VALOR TOTAL"])
    cpu.append(["1.0", "MÃO DE OBRA"])
    cpu.append(["MOD", "Ajudante de montagem", "h", 10, 4, 40])
    cpu.append(["TOTAL 1.0", 40])
    cpu.append(["2.0", "EQUIPAMENTOS"])
    cpu.append(["TOTAL 2.0", 30])
    cpu.append(["3.0", "MATERIAIS"])
    cpu.append(["TOTAL 3.0", 20])
    cpu.append(["CUSTO DIRETO (A+B+C)", 90])
    cpu.append(["BDI", 10])
    cpu.append(["TOTAL GERAL", total_geral])

    workbook.save(path)
    return path
