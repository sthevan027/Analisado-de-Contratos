import unicodedata
from dataclasses import replace
from decimal import Decimal
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook

from domain.models import ComposicaoCusto, ContratoNormalizado, ItemOrcamento, StatusCPU

_CENTAVOS = Decimal("0.01")
_ORDEM_GRUPOS = ("1.0", "2.0", "3.0")
_NOME_GRUPO = {
    "1.0": "Mão de Obra",
    "2.0": "Equipamentos",
    "3.0": "Materiais",
}
_ROTULOS_CONHECIDOS = {"ITEM:", "TOTAL 1.0", "TOTAL 2.0", "TOTAL 3.0", "TOTAL GERAL"}


def _normalizar(valor: object) -> str:
    texto = "" if valor is None else str(valor)
    sem_acento = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return sem_acento.strip().upper()


def _para_decimal(valor: object) -> Decimal:
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))
    return Decimal("0")


def _selecionar_aba(nomes_abas: list[str]) -> str:
    for nome in nomes_abas:
        if _normalizar(nome) == "CPU":
            return nome
    raise ValueError("Nenhuma aba CPU encontrada na planilha")


def _encontrar_rotulo(linha: tuple) -> tuple[int, str] | None:
    for coluna, valor in enumerate(linha):
        if valor is None:
            continue
        texto = _normalizar(valor)
        if texto in _ROTULOS_CONHECIDOS or texto.startswith("CUSTO DIRETO"):
            return coluna, texto
    return None


def _valor_apos(linha: tuple, coluna: int) -> object:
    for proxima in range(coluna + 1, len(linha)):
        if linha[proxima] is not None:
            return linha[proxima]
    return None


def _finalizar_bloco(
    codigo: str,
    totais_diretos: dict[str, Decimal],
    custo_direto: Decimal | None,
    total_geral: Decimal | None,
) -> tuple[ComposicaoCusto, ...] | None:
    if custo_direto is None or total_geral is None:
        return None
    if not all(grupo in totais_diretos for grupo in _ORDEM_GRUPOS):
        return None

    if custo_direto == 0:
        return tuple(
            ComposicaoCusto(grupo=_NOME_GRUPO[grupo], subgrupo="", nome=_NOME_GRUPO[grupo], valor_planejado=Decimal("0"))
            for grupo in _ORDEM_GRUPOS
        )

    fator = total_geral / custo_direto
    valores: dict[str, Decimal] = {}
    for grupo in _ORDEM_GRUPOS[:-1]:
        valores[grupo] = (totais_diretos[grupo] * fator).quantize(_CENTAVOS)
    ultimo_grupo = _ORDEM_GRUPOS[-1]
    valores[ultimo_grupo] = total_geral - sum(valores.values(), Decimal("0"))

    return tuple(
        ComposicaoCusto(
            grupo=_NOME_GRUPO[grupo],
            subgrupo="",
            nome=_NOME_GRUPO[grupo],
            valor_planejado=valores[grupo],
        )
        for grupo in _ORDEM_GRUPOS
    )


def ler_cpu(path: Path) -> dict[str, tuple[ComposicaoCusto, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    aba = _selecionar_aba(workbook.sheetnames)
    planilha = workbook[aba]

    resultado: dict[str, tuple[ComposicaoCusto, ...]] = {}
    codigo_atual: str | None = None
    totais_diretos: dict[str, Decimal] = {}
    custo_direto: Decimal | None = None
    total_geral: Decimal | None = None

    def finalizar_atual() -> None:
        if codigo_atual is None:
            return
        composicao = _finalizar_bloco(codigo_atual, totais_diretos, custo_direto, total_geral)
        if composicao is not None:
            resultado[codigo_atual] = composicao

    for linha in planilha.iter_rows(values_only=True):
        achado = _encontrar_rotulo(linha)
        if achado is None:
            continue
        coluna, rotulo = achado

        if rotulo == "ITEM:":
            finalizar_atual()
            codigo_bruto = _valor_apos(linha, coluna)
            codigo_atual = str(codigo_bruto).strip() if codigo_bruto is not None else None
            totais_diretos = {}
            custo_direto = None
            total_geral = None
            continue

        if codigo_atual is None:
            continue

        valor_bruto = _valor_apos(linha, coluna)
        if rotulo == "TOTAL 1.0":
            totais_diretos["1.0"] = _para_decimal(valor_bruto)
        elif rotulo == "TOTAL 2.0":
            totais_diretos["2.0"] = _para_decimal(valor_bruto)
        elif rotulo == "TOTAL 3.0":
            totais_diretos["3.0"] = _para_decimal(valor_bruto)
        elif rotulo.startswith("CUSTO DIRETO"):
            custo_direto = _para_decimal(valor_bruto)
        elif rotulo == "TOTAL GERAL":
            total_geral = _para_decimal(valor_bruto)

    finalizar_atual()
    return resultado


def _escalar_composicao(
    composicao_unitaria: tuple[ComposicaoCusto, ...], quantidade: Decimal, total_alvo: Decimal
) -> tuple[ComposicaoCusto, ...]:
    escalada: list[ComposicaoCusto] = []
    acumulado = Decimal("0")
    for indice, parte in enumerate(composicao_unitaria):
        if indice < len(composicao_unitaria) - 1:
            valor = (parte.valor_planejado * quantidade).quantize(_CENTAVOS)
            acumulado += valor
        else:
            valor = total_alvo - acumulado
        escalada.append(replace(parte, valor_planejado=valor))
    return tuple(escalada)


def aplicar_cpu(
    contrato: ContratoNormalizado, cpu: Mapping[str, tuple[ComposicaoCusto, ...]]
) -> ContratoNormalizado:
    itens_atualizados: list[ItemOrcamento] = []
    avisos = list(contrato.avisos)

    for item in contrato.itens:
        composicao_unitaria = cpu.get(item.codigo)
        if not composicao_unitaria:
            itens_atualizados.append(item)
            continue

        valor_unitario_cpu = sum((parte.valor_planejado for parte in composicao_unitaria), Decimal("0"))
        diferenca = abs(valor_unitario_cpu - item.valor_unitario)

        if diferenca <= _CENTAVOS:
            status = StatusCPU.VALIDADO_CPU
            total_alvo = item.valor_planejado
        else:
            status = StatusCPU.DIVERGENTE_CPU
            total_alvo = (valor_unitario_cpu * item.quantidade).quantize(_CENTAVOS)
            avisos.append(
                f"Item {item.codigo}: valor unitário do CPU ({valor_unitario_cpu}) diverge do valor "
                f"unitário QQP ({item.valor_unitario})"
            )

        composicao = _escalar_composicao(composicao_unitaria, item.quantidade, total_alvo)
        itens_atualizados.append(replace(item, composicao=composicao, status_cpu=status))

    return ContratoNormalizado(formato=contrato.formato, itens=tuple(itens_atualizados), avisos=tuple(avisos))
