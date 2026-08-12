import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from domain.models import ContratoNormalizado, FormatoContrato, ItemOrcamento

_PADRAO_LD = re.compile(r"^LD\d+$")


def _normalizar(valor: object) -> str:
    texto = "" if valor is None else str(valor)
    sem_acento = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return sem_acento.strip().upper()


@dataclass(frozen=True)
class _Cabecalho:
    item_col: int
    descricao_col: int
    unidade_col: int
    quantidade_col: int
    valor_unitario_col: int | None
    valor_total_col: int | None
    data_start: int


def _selecionar_aba(nomes_abas: list[str]) -> str:
    for nome in nomes_abas:
        if _normalizar(nome) == "QQP":
            return nome
    for nome in nomes_abas:
        if _PADRAO_LD.match(_normalizar(nome)):
            return nome
    raise ValueError("Nenhuma aba QQP ou LD<número> encontrada na planilha")


def _localizar_cabecalho(linhas: list[tuple]) -> _Cabecalho:
    for indice, linha in enumerate(linhas):
        normalizados = {coluna: _normalizar(valor) for coluna, valor in enumerate(linha) if valor is not None}
        tem_item = any(texto == "ITEM" for texto in normalizados.values())
        tem_descricao = any("DESCRICAO" in texto for texto in normalizados.values())
        tem_unidade = any(texto.startswith("UNID") for texto in normalizados.values())
        tem_quantidade = any(texto.startswith("QUANT") for texto in normalizados.values())
        if not (tem_item and tem_descricao and tem_unidade and tem_quantidade):
            continue

        item_col = next(coluna for coluna, texto in normalizados.items() if texto == "ITEM")
        descricao_col = next(coluna for coluna, texto in normalizados.items() if "DESCRICAO" in texto)
        unidade_col = next(coluna for coluna, texto in normalizados.items() if texto.startswith("UNID"))
        quantidade_col = next(coluna for coluna, texto in normalizados.items() if texto.startswith("QUANT"))

        valor_unitario_col: int | None = None
        valor_total_col: int | None = None
        ultima_linha_valor = indice
        for offset in (0, 1, 2):
            linha_seguinte_idx = indice + offset
            if linha_seguinte_idx >= len(linhas):
                continue
            candidatos = {
                coluna: _normalizar(valor)
                for coluna, valor in enumerate(linhas[linha_seguinte_idx])
                if valor is not None
            }
            if valor_unitario_col is None:
                encontrado = next((coluna for coluna, texto in candidatos.items() if texto.startswith("UNITARIO")), None)
                if encontrado is not None:
                    valor_unitario_col = encontrado
                    ultima_linha_valor = max(ultima_linha_valor, linha_seguinte_idx)
            if valor_total_col is None:
                encontrado = next((coluna for coluna, texto in candidatos.items() if texto == "TOTAL"), None)
                if encontrado is not None:
                    valor_total_col = encontrado
                    ultima_linha_valor = max(ultima_linha_valor, linha_seguinte_idx)

        return _Cabecalho(
            item_col=item_col,
            descricao_col=descricao_col,
            unidade_col=unidade_col,
            quantidade_col=quantidade_col,
            valor_unitario_col=valor_unitario_col,
            valor_total_col=valor_total_col,
            data_start=ultima_linha_valor + 2,
        )

    raise ValueError("Cabeçalho da planilha QQP não foi encontrado nas primeiras 30 linhas")


def _celula(linha: tuple, coluna: int | None) -> object:
    if coluna is None or coluna >= len(linha):
        return None
    return linha[coluna]


def ler_qqp(path: Path) -> ContratoNormalizado:
    workbook = load_workbook(path, read_only=True, data_only=True)
    aba = _selecionar_aba(workbook.sheetnames)
    planilha = workbook[aba]
    linhas = list(planilha.iter_rows(min_row=1, max_row=30, values_only=True))
    cabecalho = _localizar_cabecalho(linhas)

    itens: list[ItemOrcamento] = []
    grupo_atual = "Sem grupo informado"

    for linha in planilha.iter_rows(min_row=cabecalho.data_start, values_only=True):
        codigo_bruto = _celula(linha, cabecalho.item_col)
        if codigo_bruto is None:
            continue
        codigo = str(codigo_bruto).strip()
        if not codigo:
            continue

        descricao_bruta = _celula(linha, cabecalho.descricao_col)
        descricao = str(descricao_bruta).strip() if descricao_bruta is not None else ""

        if "." not in codigo:
            grupo_atual = descricao or "Sem grupo informado"
            continue

        unidade_bruta = _celula(linha, cabecalho.unidade_col)
        unidade = str(unidade_bruta).strip() if isinstance(unidade_bruta, str) else ""
        quantidade_bruta = _celula(linha, cabecalho.quantidade_col)
        if not unidade or not isinstance(quantidade_bruta, (int, float)):
            continue

        valor_unitario_bruto = _celula(linha, cabecalho.valor_unitario_col)
        valor_total_bruto = _celula(linha, cabecalho.valor_total_col)

        quantidade = Decimal(str(quantidade_bruta))
        valor_unitario = (
            Decimal(str(valor_unitario_bruto)) if isinstance(valor_unitario_bruto, (int, float)) else Decimal("0")
        )
        if isinstance(valor_total_bruto, (int, float)):
            valor_planejado = Decimal(str(valor_total_bruto))
        else:
            valor_planejado = quantidade * valor_unitario

        itens.append(
            ItemOrcamento(
                codigo=codigo,
                descricao=descricao,
                unidade=unidade,
                quantidade=quantidade,
                valor_unitario=valor_unitario,
                valor_planejado=valor_planejado,
                area=grupo_atual,
            )
        )

    return ContratoNormalizado(formato=FormatoContrato.QQP, itens=tuple(itens))
